import io
import json
import datetime
from datetime import timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count, Case, When, IntegerField
from django.db.models.functions import TruncDate, TruncHour
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Customer, Gift
from .forms import CustomerRegistrationForm
from .decorators import admin_required


@login_required
def register(request):
    if request.method == 'POST':
        form = CustomerRegistrationForm(request.POST)
        if form.is_valid():
            customer = form.save(commit=False)
            customer.has_played = True  # registration = entered the game
            customer.save()
            return redirect('registration_success', pk=customer.pk)
        else:
            messages.error(request, 'Please fix the errors below and try again.')
    else:
        form = CustomerRegistrationForm()

    gifts = Gift.objects.filter(is_active=True)
    return render(request, 'customers/register.html', {
        'form': form,
        'gifts': gifts,
        'consolation_gift': Customer.CONSOLATION_GIFT,
    })


@login_required
def registration_success(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if not customer.has_played:
        customer.has_played = True
        customer.save(update_fields=['has_played'])
    return render(request, 'customers/success.html', {'customer': customer})


@admin_required
def dashboard(request):
    from datetime import date as date_type
    search_query = request.GET.get('q', '').strip()
    date_filter  = request.GET.get('date', '').strip()

    today = timezone.localtime(timezone.now()).date()
    yesterday = today - timedelta(days=1)

    # Resolve selected date range or single date
    from_date = None
    to_date = None
    if date_filter:
        if ' to ' in date_filter:
            try:
                parts = date_filter.split(' to ')
                from_date = date_type.fromisoformat(parts[0].strip())
                to_date = date_type.fromisoformat(parts[1].strip())
            except (ValueError, IndexError):
                date_filter = ''
        else:
            try:
                from_date = date_type.fromisoformat(date_filter.strip())
                to_date = from_date
            except ValueError:
                date_filter = ''

    if not from_date or not to_date:
        from_date = today
        to_date = today

    # Format the display date/range text
    if from_date == to_date:
        date_display = from_date.strftime('%d %b %Y')
    else:
        date_display = f"{from_date.strftime('%d %b %Y')} to {to_date.strftime('%d %b %Y')}"

    # Build local time zone aware datetimes for filtering (avoiding database CONVERT_TZ reliance)
    today_start = timezone.make_aware(datetime.datetime.combine(today, datetime.time.min))
    today_end = timezone.make_aware(datetime.datetime.combine(today, datetime.time.max))

    from_datetime = timezone.make_aware(datetime.datetime.combine(from_date, datetime.time.min))
    to_datetime = timezone.make_aware(datetime.datetime.combine(to_date, datetime.time.max))

    # Single aggregation query for all stat card counts
    agg = Customer.objects.aggregate(
        total_all=Count('id'),
        played_all=Count(Case(When(has_played=True, then=1), output_field=IntegerField())),
        today=Count(Case(When(registered_at__gte=today_start, registered_at__lte=today_end, then=1), output_field=IntegerField())),
        filtered_total=Count(Case(When(registered_at__gte=from_datetime, registered_at__lte=to_datetime, then=1), output_field=IntegerField())),
        filtered_played=Count(Case(When(registered_at__gte=from_datetime, registered_at__lte=to_datetime, has_played=True, then=1), output_field=IntegerField())),
        winners_all=Count(Case(When(game_result=Customer.RESULT_WIN, then=1), output_field=IntegerField())),
        losers_all=Count(Case(When(game_result=Customer.RESULT_LOSS, then=1), output_field=IntegerField())),
        filtered_win=Count(Case(When(registered_at__gte=from_datetime, registered_at__lte=to_datetime, game_result=Customer.RESULT_WIN, then=1), output_field=IntegerField())),
        filtered_loss=Count(Case(When(registered_at__gte=from_datetime, registered_at__lte=to_datetime, game_result=Customer.RESULT_LOSS, then=1), output_field=IntegerField())),
    )

    if date_filter:
        total_count  = agg['filtered_total']
        played_count = agg['filtered_played']
        winners_count = agg['filtered_win']
        losers_count  = agg['filtered_loss']
    else:
        total_count  = agg['total_all']
        played_count = agg['played_all']
        winners_count = agg['winners_all']
        losers_count  = agg['losers_all']
    today_count      = agg['today']

    # Bar chart: 7-day window for default, or exact selected range/date when filtered
    if date_filter:
        start_date = from_date
        end_date = to_date
    else:
        # Default: 7-day trend ending on current day
        start_date = from_date - timedelta(days=6)
        end_date = from_date

    start_datetime = timezone.make_aware(datetime.datetime.combine(start_date, datetime.time.min))
    end_datetime = timezone.make_aware(datetime.datetime.combine(end_date, datetime.time.max))

    daily_qs = (
        Customer.objects
        .filter(registered_at__gte=start_datetime, registered_at__lte=end_datetime)
        .values_list('registered_at', flat=True)
    )
    daily_map = {}
    for dt in daily_qs:
        if dt:
            local_dt = timezone.localtime(dt)
            label = local_dt.strftime('%d %b')
            daily_map[label] = daily_map.get(label, 0) + 1

    chart_labels = []
    chart_data = []
    num_days = (end_date - start_date).days + 1
    for i in range(num_days):
        day = start_date + timedelta(days=i)
        label = day.strftime('%d %b')
        chart_labels.append(label)
        chart_data.append(daily_map.get(label, 0))

    # Hourly chart: scoped to selected date range, 9 AM – 9 PM window
    EVENT_START = 9
    EVENT_END   = 21
    
    hourly_qs = (
        Customer.objects
        .filter(registered_at__gte=from_datetime, registered_at__lte=to_datetime)
        .values_list('registered_at', flat=True)
    )
    hourly_map = {}
    for dt in hourly_qs:
        if dt:
            local_dt = timezone.localtime(dt)
            hour = local_dt.hour
            hourly_map[hour] = hourly_map.get(hour, 0) + 1

    hourly_labels = [f'{h:02d}:00' for h in range(EVENT_START, EVENT_END + 1)]
    hourly_data   = [hourly_map.get(h, 0) for h in range(EVENT_START, EVENT_END + 1)]

    # All unique dates that have registrations (for calendar highlighting)
    all_registrations = Customer.objects.values_list('registered_at', flat=True)
    registered_dates = {timezone.localtime(dt).date() for dt in all_registrations if dt}
    registered_dates_json = json.dumps([d.isoformat() for d in registered_dates])

    # Gift winning counts breakdown: all-time by default, scoped to the selected
    # date range only when a date filter is actually applied.
    gift_win_filter = Q(winners__game_result=Customer.RESULT_WIN)
    consolation_filter = Q(game_result=Customer.RESULT_LOSS)
    if date_filter:
        gift_win_filter &= Q(winners__registered_at__gte=from_datetime, winners__registered_at__lte=to_datetime)
        consolation_filter &= Q(registered_at__gte=from_datetime, registered_at__lte=to_datetime)

    gifts_qs = Gift.objects.filter(is_active=True).annotate(
        win_count=Count('winners', filter=gift_win_filter)
    ).order_by('-win_count')

    gift_labels = [g.name for g in gifts_qs]
    gift_data = [g.win_count for g in gifts_qs]

    # Append the complimentary/consolation gift given to all Loss participants
    consolation_count = Customer.objects.filter(consolation_filter).count()
    gift_labels.append(f'{Customer.CONSOLATION_GIFT} (Complimentary)')
    gift_data.append(consolation_count)

    context = {
        'search_query':  search_query,
        'date_filter':   date_filter,
        'date_display':  date_display,
        'from_date':     from_date,
        'to_date':       to_date,
        'active_date':   from_date,
        'today_iso':     today.isoformat(),
        'yesterday_iso': yesterday.isoformat(),
        'total_count':   total_count,
        'played_count':  played_count,
        'today_count':   today_count,
        'winners_count': winners_count,
        'losers_count':  losers_count,
        'chart_labels':  json.dumps(chart_labels),
        'chart_data':    json.dumps(chart_data),
        'hourly_labels': json.dumps(hourly_labels),
        'hourly_data':   json.dumps(hourly_data),
        'gift_labels':   json.dumps(gift_labels),
        'gift_data':     json.dumps(gift_data),
        'registered_dates_json': registered_dates_json,
    }
    return render(request, 'customers/dashboard.html', context)


@admin_required
def report(request):
    from datetime import date as date_type
    from_date_str = request.GET.get('from_date', '').strip()
    to_date_str   = request.GET.get('to_date', '').strip()
    search_query  = request.GET.get('q', '').strip()
    result_filter = request.GET.get('result', '').strip()
    gift_filter   = request.GET.get('gift', '').strip()

    customers = Customer.objects.all()

    from_date = None
    to_date   = None

    # Timezone-aware datetime range instead of __date lookups, which rely on the
    # database's CONVERT_TZ (breaks silently on MySQL without timezone tables loaded).
    if from_date_str:
        try:
            from_date = date_type.fromisoformat(from_date_str)
            from_datetime = timezone.make_aware(datetime.datetime.combine(from_date, datetime.time.min))
            customers = customers.filter(registered_at__gte=from_datetime)
        except ValueError:
            from_date_str = ''

    if to_date_str:
        try:
            to_date = date_type.fromisoformat(to_date_str)
            to_datetime = timezone.make_aware(datetime.datetime.combine(to_date, datetime.time.max))
            customers = customers.filter(registered_at__lte=to_datetime)
        except ValueError:
            to_date_str = ''

    if search_query:
        customers = customers.filter(
            Q(name__icontains=search_query) |
            Q(mobile_number__icontains=search_query) |
            Q(bill_number__icontains=search_query)
        )

    if result_filter in dict(Customer.RESULT_CHOICES):
        customers = customers.filter(game_result=result_filter)
    else:
        result_filter = ''

    selected_gift = None
    if gift_filter:
        if gift_filter.isdigit():
            selected_gift = Gift.objects.filter(pk=gift_filter).first()
            if selected_gift:
                customers = customers.filter(won_gift_id=gift_filter)
            else:
                gift_filter = ''
        else:
            gift_filter = ''

    total_filtered = customers.count()

    paginator   = Paginator(customers, 50)
    page_number = request.GET.get('page')
    page_obj    = paginator.get_page(page_number)

    # Assign chronological serial number (oldest gets 1, newest gets highest number)
    total_count = paginator.count
    start_idx = page_obj.start_index()
    for idx, customer in enumerate(page_obj.object_list):
        customer.serial_number = total_count - (start_idx - 1) - idx

    # Shared query string (without 'page') reused by export link and pagination links
    filter_params = {
        'from_date': from_date_str,
        'to_date':   to_date_str,
        'q':         search_query,
        'result':    result_filter,
        'gift':      gift_filter,
    }
    from urllib.parse import quote
    filter_qs = '&'.join(f'{k}={quote(v)}' for k, v in filter_params.items() if v)

    context = {
        'page_obj':      page_obj,
        'search_query':  search_query,
        'from_date_str': from_date_str,
        'to_date_str':   to_date_str,
        'filter_qs':     filter_qs,
        'from_date':     from_date,
        'to_date':       to_date,
        'total_filtered': total_filtered,
        'result_filter': result_filter,
        'gift_filter':   gift_filter,
        'selected_gift': selected_gift,
        'gifts':         Gift.objects.all(),
    }
    return render(request, 'customers/report.html', context)


@admin_required
def export_report_excel(request):
    from datetime import date as date_type
    from_date_str = request.GET.get('from_date', '').strip()
    to_date_str   = request.GET.get('to_date', '').strip()
    search_query  = request.GET.get('q', '').strip()
    result_filter = request.GET.get('result', '').strip()
    gift_filter   = request.GET.get('gift', '').strip()

    customers = Customer.objects.all()

    if from_date_str:
        try:
            from_date = date_type.fromisoformat(from_date_str)
            from_datetime = timezone.make_aware(datetime.datetime.combine(from_date, datetime.time.min))
            customers = customers.filter(registered_at__gte=from_datetime)
        except ValueError:
            pass

    if to_date_str:
        try:
            to_date = date_type.fromisoformat(to_date_str)
            to_datetime = timezone.make_aware(datetime.datetime.combine(to_date, datetime.time.max))
            customers = customers.filter(registered_at__lte=to_datetime)
        except ValueError:
            pass

    if search_query:
        customers = customers.filter(
            Q(name__icontains=search_query) |
            Q(mobile_number__icontains=search_query) |
            Q(bill_number__icontains=search_query)
        )

    if result_filter in dict(Customer.RESULT_CHOICES):
        customers = customers.filter(game_result=result_filter)

    if gift_filter.isdigit():
        customers = customers.filter(won_gift_id=gift_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Customer Report'

    # Header row styling
    header_fill   = PatternFill('solid', fgColor='E63946')
    header_font   = Font(bold=True, color='FFFFFF', size=11)
    header_align  = Alignment(horizontal='center', vertical='center')
    thin_border   = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    headers = ['#', 'Name', 'Mobile Number', 'Aadhar Number', 'Bill Number', 'Status', 'Game Result', 'Gift', 'Registered Date', 'Registered Time']
    col_widths = [6, 28, 18, 20, 18, 12, 12, 32, 18, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 24

    # Alternating row fill
    alt_fill = PatternFill('solid', fgColor='FFF5F5')
    data_font = Font(size=10)
    center_align = Alignment(horizontal='center', vertical='center')

    total_count = customers.count()
    for idx, customer in enumerate(customers.iterator()):
        row_idx = idx + 2
        row_fill = alt_fill if row_idx % 2 == 0 else None
        local_time = timezone.localtime(customer.registered_at)
        row_data = [
            total_count - idx,
            customer.name,
            customer.mobile_number,
            customer.aadhar_number,
            customer.bill_number,
            'Played' if customer.has_played else 'Pending',
            customer.get_game_result_display(),
            customer.gift_display or '—',
            local_time.strftime('%d-%m-%Y'),
            local_time.strftime('%H:%M'),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font   = data_font
            cell.border = thin_border
            cell.alignment = center_align if col_idx in (1, 6, 7, 9, 10) else Alignment(vertical='center')
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[row_idx].height = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Build filename
    parts = ['customer_report']
    if from_date_str:
        parts.append(f'from_{from_date_str}')
    if to_date_str:
        parts.append(f'to_{to_date_str}')
    filename = '_'.join(parts) + '.xlsx'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def check_duplicate(request):
    field = request.GET.get('field')
    value = request.GET.get('value', '').strip()

    if not field or not value:
        return JsonResponse({'exists': False})

    exists = False
    if field == 'mobile_number':
        exists = Customer.objects.filter(mobile_number=value).exists()
    elif field == 'aadhar_number':
        exists = Customer.objects.filter(aadhar_number=value).exists()
    elif field == 'bill_number':
        exists = Customer.objects.filter(bill_number=value.upper()).exists()

    return JsonResponse({'exists': exists})


from django.contrib.auth.models import User
from .forms import UserForm

@admin_required
def user_list(request):
    users = User.objects.all().order_by('-is_staff', 'username')
    return render(request, 'customers/user_list.html', {'users': users})


@admin_required
def user_create(request):
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'User created successfully.')
            return redirect('user_list')
    else:
        form = UserForm()
    return render(request, 'customers/user_form.html', {'form': form, 'title': 'Create User'})


@admin_required
def user_edit(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'User updated successfully.')
            return redirect('user_list')
    else:
        form = UserForm(instance=user)
    return render(request, 'customers/user_form.html', {'form': form, 'title': 'Edit User', 'editing': True})


@admin_required
def user_delete(request, pk):
    user = get_object_or_404(User, pk=pk)
    if user == request.user:
        messages.error(request, 'You cannot delete your own user account.')
        return redirect('user_list')
    
    if request.method == 'POST':
        user.delete()
        messages.success(request, 'User deleted successfully.')
        return redirect('user_list')
    
    return render(request, 'customers/user_confirm_delete.html', {'target_user': user})
